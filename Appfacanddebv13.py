
import streamlit as st
import pandas as pd
import re
import csv
from io import BytesIO
import zipfile
import traceback
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image as XLImage
from PIL import Image


# Columnas a eliminar completamente
columns_to_drop = [
    'FECHA REND', 'IMPORTE REND.HC', 'ALIC.IVA', 'QUIEN FAC.', 'HORA',
    'PANTALLA', 'ADMIS', 'TIPO DE MARCA', 'PROTOCOLO 1', 'PROTOCOLO 2',
    'PROTOCOLO 3', 'PROTOCOLO 4', 'PROTOCOLO 5', 'COD.MA'
]

# Orden deseado de columnas
column_order = [
    'H.CLINICA', 'HC UNICA', 'APELLIDO Y NOMBRE', 'AFILIADO', 'PERIODO',
    'COD.OBRA', 'COBERTURA', 'PLAN', 'NRO.FACTURA', 'FECHA PRES',
    'TIP.NOM', 'COD.NOM', 'PRESTACION', 'CANTID.', 'IMPORTE UNIT.',
    'IMPORTE PREST.', 'ORIGEN'
]

# Columnas que deben convertirse a numérico
numeric_columns = [
    'H.CLINICA', 'HC UNICA', 'AFILIADO', 'TIP.NOM',
    'COD.NOM', 'CANTID.', 'IMPORTE UNIT.',
    'COD.OBRA.', 'IMPORTE PREST.',
]

def ensure_pipe_at_end(file):
    content = file.read().decode('utf-8')
    lines = content.splitlines()
    if lines and not lines[0].endswith('|'):
        lines[0] += '|'
    corrected_content = '\n'.join(lines)
    return BytesIO(corrected_content.encode('utf-8'))

def detectar_delimitador(file_like, default='|'):
    file_like.seek(0)
    sample = file_like.read(2048).decode('utf-8', errors='ignore')
    try:
        dialect = csv.Sniffer().sniff(sample)
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = default
    file_like.seek(0)
    return delimiter

def leer_txt_a_dataframe(file):
    corrected_file = ensure_pipe_at_end(file)
    delimiter = detectar_delimitador(corrected_file)
    corrected_file.seek(0)
    df = pd.read_csv(corrected_file, delimiter=delimiter, dtype=str)
    if df.empty:
        raise ValueError("El archivo fue leído pero no contiene datos.")
    return df

def clean_and_format_dataframe(df):
    df = df.drop(columns=[col for col in columns_to_drop if col in df.columns], errors='ignore')
    existing_columns = [col for col in column_order if col in df.columns]
    df = df[existing_columns + [col for col in df.columns if col not in existing_columns]]
    for col in numeric_columns:
        if col in df.columns:
            if pd.api.types.is_string_dtype(df[col]):
                df[col] = df[col].str.replace(',', '.', regex=False)
            df[col] = pd.to_numeric(df[col], errors='coerce')
    return df


def generate_zip_with_summary(df, folder_base, modo_operacion, logo_bytes):
    zip_buffer = BytesIO()
    safe_base = re.sub(r'\W+', '_', folder_base.strip()) or "Facturas"

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
        grouped = df.groupby(['COBERTURA', 'NRO.FACTURA'])

        for (cobertura, factura), group in grouped:
            safe_cobertura = re.sub(r'\W+', '', str(cobertura))[:20]
            safe_factura = re.sub(r'\W+', '', str(factura))[:20]
            filename = f"{safe_base}/{safe_cobertura}/Factura_{safe_factura}_{safe_cobertura}.xlsx"

            if modo_operacion == "Débitos":
                columnas_deseadas = [
                    "APELLIDO Y NOMBRE", "COD.NOM", "PRESTACION", "FECHA PRES",
                    "CANTID.", "IMPORTE UNIT.", "IMPORTE PREST."
                ]
                
                group = clean_and_format_dataframe(group)  # ← Esta línea es nueva
                group = group[[col for col in columnas_deseadas if col in group.columns]]

                wb = load_workbook("Plantilla_Débitos_1.xlsx")
                ws = wb["Debitos"]
     
               
                # Buscar la tabla que contenga 'debitos' en su nombre
                tabla = None
                for t_name in ws.tables:
                    if "debitos" in t_name.lower():
                        tabla = ws.tables[t_name]
                        break

                # Validar que se encontró la tabla
                if not tabla:
                    st.warning(f"No se encontró una tabla que contenga 'debitos' en su nombre. Tablas disponibles: {list(ws.tables.keys())}")
                   

                # Insertar logo en A1 si se proporcionó 
                if logo_bytes:
                    try:
                        logo_img = Image.open(BytesIO(logo_bytes))
                        logo_img.save("temp_logo.png")  # Guardar temporalmente
                        xl_logo = XLImage("temp_logo.png")
                        xl_logo.anchor = 'A1'
                        ws.add_image(xl_logo)
                    except Exception as e:
                        st.warning(f"No se pudo insertar el logo: {e}")
    
                # Insertar encabezado
                encabezado = f"REFACTURACIÓN Fc {factura} - {cobertura}"
                ws["B1"] = encabezado


                # Insertar datos desde la fila 3, solo columnas A-G
                start_row = 3
                for i, row in group.iterrows():
                    for j, col in enumerate(columnas_deseadas[:7], start=1):  # A-G
                        if col in row:
                            ws.cell(row=start_row, column=j, value=row[col])
                    start_row += 1
                    
          
                # Crear validación de datos para SECTOR (columna K) con lista fija
                dv_sector = DataValidation(type="list", formula1='"ADMIN,MEDICO,COMERCIAL"', allow_blank=True)

                # Agregar la validación a la hoja
                ws.add_data_validation(dv_sector)

                # Aplicar la validación a un rango amplio en la columna K (por ejemplo, de la fila 3 a la 200)
                dv_sector.add("K3:K200")

                
                # Actualizar rango de la tabla (recrear tabla con nuevo rango)
                if tabla:
                    try:
                        new_ref = f"{ws.cell(row=2, column=1).coordinate}:{ws.cell(row=start_row - 1, column=20).coordinate}"
                        new_table = Table(displayName=tabla.displayName, ref=new_ref)
                        new_table.tableStyleInfo = tabla.tableStyleInfo
                        del ws.tables[tabla.displayName]
                        ws.add_table(new_table)
                    except Exception as e:
                        st.warning(f"No se pudo actualizar el rango de la tabla: {e}")

               
                final_buffer = BytesIO()
                wb.save(final_buffer)
                final_buffer.seek(0)
                zipf.writestr(filename, final_buffer.read())
                
            else:
                group = clean_and_format_dataframe(group)
                excel_buffer = BytesIO()
                group.to_excel(excel_buffer, index=False, engine='openpyxl')
                excel_buffer.seek(0)
                zipf.writestr(filename, excel_buffer.read())

        # Crear resumen
        df['IMPORTE PREST.'] = pd.to_numeric(df['IMPORTE PREST.'], errors='coerce').fillna(0)
        summary_df = (
            df.groupby(['COBERTURA', 'NRO.FACTURA', 'APELLIDO Y NOMBRE'], as_index=False)['IMPORTE PREST.']
            .sum()
        )
        summary_buffer = BytesIO()
        summary_df.to_excel(summary_buffer, index=False, engine='openpyxl')
        summary_buffer.seek(0)
        zipf.writestr(f"{safe_base}/resumen_facturas.xlsx", summary_buffer.read())

    zip_buffer.seek(0)
    return zip_buffer

 
def process_file(file, folder_base, modo_operacion, logo_bytes, selected_facturas=None):
    try:
        try:
            df = leer_txt_a_dataframe(file)
        except Exception as e:
            st.error(f"Error al leer el archivo {file.name}: {e}")
            return

        df.columns = df.columns.str.strip()
        required_columns = ['NRO.FACTURA', 'COBERTURA']
        missing = [col for col in required_columns if col not in df.columns]
        if missing:
            st.error(f"Faltan las siguientes columnas requeridas: {', '.join(missing)}")
            return

        # Limpieza de espacios en blanco
        for col in df.select_dtypes(include='object'):
            df[col] = df[col].map(lambda x: x.strip() if isinstance(x, str) else x)

        df.dropna(how='all', inplace=True)
        df.sort_values(by='NRO.FACTURA', inplace=True)

        # Filtrado por facturas seleccionadas
        if modo_operacion == "Débitos" and selected_facturas:
            selected_facturas = [str(f).strip() for f in selected_facturas]
            df = df[df['NRO.FACTURA'].astype(str).str.strip().isin(selected_facturas)]

        df_clean = clean_and_format_dataframe(df)
        
        unique_invoices = df_clean['NRO.FACTURA'].nunique()
        st.info(f"Se generarán {unique_invoices} archivos únicos por número de factura.")

        zip_output = generate_zip_with_summary(df, folder_base, modo_operacion, logo_bytes)
        st.success("Archivo convertido y listo para descargar.")
        st.download_button("📦 Descargar ZIP con facturas y resumen", data=zip_output, file_name="facturas_por_cobertura.zip", mime="application/zip")

    except Exception as e:
        st.error(f"Ocurrió un error: {e}")
        st.text(traceback.format_exc())

# Interfaz de usuario
st.title("📄 Convertidor TXT a Excel con separación por COBERTURA y resumen")

uploaded_files = st.file_uploader("Selecciona uno o más archivos .txt para convertir a Excel", type="txt", accept_multiple_files=True)
folder_base = st.text_input("📁 Nombre de la carpeta raíz para los archivos generados", value="Facturas")
modo_operacion = st.selectbox("Selecciona el tipo de operación", ["Facturación", "Débitos"])

# Multiselección de facturas si es Débitos
selected_facturas = []

if uploaded_files and modo_operacion == "Débitos":
    try:
        file_bytes = uploaded_files[0].getvalue()
        file_copy = BytesIO(file_bytes)
        df_preview = leer_txt_a_dataframe(file_copy)
        df_preview.columns = df_preview.columns.str.strip()

        if 'NRO.FACTURA' in df_preview.columns:
            facturas_unicas = sorted(df_preview['NRO.FACTURA'].dropna().astype(str).unique())

            # Campo para pegar números de factura manualmente
            facturas_pegadas = st.text_area(
                "📋 O pega aquí los números de factura separados por coma, espacio o salto de línea",
                placeholder="Ej: 12345, 67890\n11223"
            )

            facturas_pegadas_lista = []
            facturas_no_encontradas = []
            facturas_encontradas = [] 
            
            if facturas_pegadas:
                import re
                facturas_pegadas_lista = re.split(r'[,\n]+', facturas_pegadas.strip())
                facturas_pegadas_lista = [f.strip() for f in facturas_pegadas_lista if f]

                
                # Buscar coincidencias por sbcadena
                for f in facturas_pegadas_lista:
                    coincidencias = []
                    for fact in facturas_unicas:
                        fact_limpio = str(fact).strip()
                        # Buscar si lo pegado está dentro del número de factura
                        if f in fact_limpio:
                            coincidencias.append(fact)
                    if coincidencias:
                        facturas_encontradas.extend(coincidencias)
                    else:
                        facturas_no_encontradas.append(f)

            if facturas_no_encontradas:
                st.warning(f"⚠️ Las siguientes entradas no se encontraron como coincidencia en los números de factura: {', '.join(facturas_no_encontradas)}")

            # Eliminar duplicados
            facturas_encontradas = sorted(set(facturas_encontradas))

            # Multiselección con valores válidos
            selected_facturas = st.multiselect(
                "🧾 Selecciona los números de factura que deseas generar",
                options=facturas_unicas,
            default=facturas_encontradas
            )

            st.caption(f"Se seleccionaron {len(selected_facturas)} factura(s).")
        else:
            st.warning("El archivo no contiene la columna 'NRO.FACTURA'.")
    except Exception as e:
        st.warning(f"No se pudo cargar la lista de facturas para seleccionar: {e}")


# Botón para procesar
if st.button("🚀 Convertir"):
    if uploaded_files:
        with st.spinner("Procesando archivos..."):
            for file in uploaded_files:
                st.subheader(f"Procesando: {file.name}")
                process_file(file, folder_base, modo_operacion, None, selected_facturas)
